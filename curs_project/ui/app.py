import os
import io
import datetime
import json
import csv
import urllib.request
from tkinter import filedialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import ttk, messagebox, Canvas
from PIL import Image, ImageTk, ImageDraw
import mysql.connector
from mysql.connector import Error
from werkzeug.security import check_password_hash, generate_password_hash
import ssl


try:
    import pandas as pd
    from docx import Document
    from docx.shared import Inches
    PANDAS_AVAILABLE = True
except ImportError:
    print("Для експорту встановіть: pip install pandas openpyxl python-docx")
    PANDAS_AVAILABLE = False

from config import DB_CONFIG, ASSETS_DIR, MAP_COORDINATES, CANVAS_WIDTH, CANVAS_HEIGHT
from database import DatabaseInitializer, safe_connect
from ui.widgets import CalendarDialog, CarCard, ModernCRUDDialog, DaysCounterWidget, ImageCarousel, MapWidget

# Создаем SSL контекст один раз
ssl_context = ssl.create_default_context()
ssl_context.check_hostname = False
ssl_context.verify_mode = ssl.CERT_NONE

# Константы для стилей
COLORS = {
    'primary': '#007bff',
    'success': '#28a745',
    'warning': '#ffc107',
    'danger': '#dc3545',
    'info': '#17a2b8',
    'light': '#f8f9fa',
    'dark': '#343a40'
}

class EnhancedAutoTrackerApp(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title("Auto Tracker Pro — Advanced Vehicle Tracking")
        self.geometry(f"{self.winfo_screenwidth()}x{self.winfo_screenheight()}")
        self.state('zoomed')
        
        # Загрузка иконки если она существует
        self._load_app_icon()
        
        if not DatabaseInitializer.initialize_database():
            messagebox.showerror("Помилка", "Не вдалося ініціалізувати базу даних")
            return

        self.conn = None
        self.current_user = None
        self.dark_mode = False
        self.current_table = None
        self.current_filters = {}
        self.view_mode = "cards"
        self.selected_purchase = None

        self.left_frame = None
        self.main_content = None

        self._build_login_ui()
        
    def _load_app_icon(self):
        """Корректная загрузка иконки для Windows / macOS / Linux"""
        try:
            png_path = os.path.join(ASSETS_DIR, "app_icon.png")
            ico_path = os.path.join(ASSETS_DIR, "app_icon.ico")

            import platform
            system = platform.system().lower()

            if system == "windows":
                # Windows поддерживает ТОЛЬКО ICO
                if os.path.exists(ico_path):
                    self.iconbitmap(ico_path)
                else:
                    print("⚠️ ICON (.ico) not found!")
            
            elif system == "darwin":  # macOS
                # На macOS .ico может не загрузиться → загружаем PNG
                if os.path.exists(png_path):
                    try:
                        img = ImageTk.PhotoImage(Image.open(png_path))
                        self.tk.call('wm', 'iconphoto', self._w, img)
                        self._app_icon_ref = img  # важно: не дать GC удалить ссылку
                    except Exception as e:
                        print(f"Error loading PNG icon: {e}")
                else:
                    print("⚠️ PNG icon not found!")
            
            else:
                # Linux: обе опции обычно работают
                if os.path.exists(png_path):
                    img = ImageTk.PhotoImage(Image.open(png_path))
                    self.tk.call('wm', 'iconphoto', self._w, img)
                    self._app_icon_ref = img
                elif os.path.exists(ico_path):
                    self.iconbitmap(ico_path)

        except Exception as e:
            print(f"⚠️ Не удалось загрузить иконку: {e}")

    def center_window(self, window, width, height):
        """Центрирование дочерних окон относительно главного"""
        window.update_idletasks()
        parent_x = self.winfo_rootx()
        parent_y = self.winfo_rooty()
        parent_width = self.winfo_width()
        parent_height = self.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        window.geometry(f"{width}x{height}+{x}+{y}")

    def _get_table_columns(self, table):
        """Получение колонок таблицы с обработкой ошибок"""
        if not self.conn:
            return []
        
        try:
            cur = self.conn.cursor()
            cur.execute(f"SHOW COLUMNS FROM `{table}`")
            columns = cur.fetchall()
            cur.close()
            return columns
        except Error as e:
            print(f"Помилка отримання колонок: {e}")
            return []

    def _get_primary_key(self, table):
        """Получение первичного ключа таблицы"""
        try:
            cur = self.conn.cursor()
            cur.execute(f"SHOW KEYS FROM `{table}` WHERE Key_name = 'PRIMARY'")
            primary_key = cur.fetchone()
            cur.close()
            
            if primary_key:
                return primary_key[4]
            return None
        except Error as e:
            print(f"Помилка отримання первинного ключа: {e}")
            return None

    def _is_foreign_key(self, table, column):
        """Проверка является ли колонка внешним ключом"""
        try:
            cur = self.conn.cursor()
            cur.execute(f"""
                SELECT CONSTRAINT_NAME 
                FROM information_schema.KEY_COLUMN_USAGE 
                WHERE TABLE_NAME = '{table}' 
                AND COLUMN_NAME = '{column}' 
                AND REFERENCED_TABLE_NAME IS NOT NULL
            """)
            result = cur.fetchone()
            cur.close()
            return result is not None
        except Error:
            return False

    def _get_foreign_key_info(self, table, column):
        """Получение информации о внешнем ключе"""
        try:
            cur = self.conn.cursor()
            cur.execute(f"""
                SELECT REFERENCED_TABLE_NAME, REFERENCED_COLUMN_NAME 
                FROM information_schema.KEY_COLUMN_USAGE 
                WHERE TABLE_NAME = '{table}' 
                AND COLUMN_NAME = '{column}' 
                AND REFERENCED_TABLE_NAME IS NOT NULL
            """)
            result = cur.fetchone()
            cur.close()
            return result if result else (None, None)
        except Error:
            return (None, None)

    def _get_foreign_key_values(self, table, column):
        """Получение значений для внешнего ключа"""
        try:
            cur = self.conn.cursor()
            cur.execute(f"SHOW COLUMNS FROM `{table}`")
            cols = cur.fetchall()
            
            display_column = None
            for col in cols:
                if col[0] in ['name', 'username', 'country_name', 'status_name', 'port_name', 'auction_name', 'location_name']:
                    display_column = col[0]
                    break
            
            if display_column and display_column != column:
                cur.execute(f"SELECT `{column}`, `{display_column}` FROM `{table}`")
                rows = cur.fetchall()
                cur.close()
                return [f"{row[0]} - {row[1]}" for row in rows if row[0] is not None]
            else:
                cur.execute(f"SELECT `{column}` FROM `{table}`")
                rows = cur.fetchall()
                cur.close()
                return [str(row[0]) for row in rows if row[0] is not None]
                
        except Error as e:
            print(f"Помилка отримання значень зовнішнього ключа: {e}")
            return []

    def _export_table_to_excel(self, table):
        """Экспорт таблицы в Excel"""
        if not PANDAS_AVAILABLE:
            messagebox.showerror("Помилка", "Для експорту в Excel встановіть: pip install pandas openpyxl")
            return
            
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not file_path:
                return
                
            query = f"SELECT * FROM `{table}`"
            cur = self.conn.cursor()
            cur.execute(query)
            rows = cur.fetchall()
            cols = [i[0] for i in cur.description]
            cur.close()
            
            df = pd.DataFrame(rows, columns=cols)
            df.to_excel(file_path, index=False, engine='openpyxl')
            
            messagebox.showinfo("Успіх", f"Експортовано в Excel: {file_path}")
            
        except Exception as e:
            messagebox.showerror("Помилка", f"Помилка експорту: {str(e)}")

    def _export_to_csv(self, table=None):
        """Экспорт таблицы в CSV"""
        if not table:
            table = self.current_table
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            try:
                query = f"SELECT * FROM `{table}`"
                cur = self.conn.cursor()
                cur.execute(query)
                rows = cur.fetchall()
                cols = [i[0] for i in cur.description]
                cur.close()

                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(cols)
                    writer.writerows(rows)

                messagebox.showinfo("Успіх", f"Експортовано в CSV: {file_path}")
            except Error as e:
                messagebox.showerror("Помилка", f"Помилка експорту: {e}")

    def _export_current_table(self, format_type):
        """Экспорт текущей таблицы в указанном формате"""
        if not hasattr(self, 'current_table') or not self.current_table:
            messagebox.showwarning("Увага", "Спочатку виберіть таблицю")
            return
            
        if format_type == 'excel':
            self._export_table_to_excel(self.current_table)
        elif format_type == 'csv':
            self._export_to_csv(self.current_table)

    def _build_login_ui(self):
        """Красивый современный интерфейс входа"""

        # Очистка окна
        for w in self.winfo_children():
            w.destroy()

        # -------------------------------------------------------
        #               Основной контейнер (2 колонки)
        # -------------------------------------------------------
        main_container = tb.Frame(self, padding=20)
        main_container.pack(fill="both", expand=True)

        # -------------------------------------------------------
        #                     ЛЕВАЯ ЧАСТЬ
        # -------------------------------------------------------
        left_side = tb.Frame(main_container, padding=40)
        left_side.pack(side="left", fill="both", expand=True)

        # ----- ЛОГО ИЗ ФАЙЛА (assets/app_icon.jpg) -----
        icon_path = os.path.join(ASSETS_DIR, "icon.png")

        try:
            pil = Image.open(icon_path)
            
            # Уменьшаем логотип, сохраняя пропорции
            pil.thumbnail((180, 180), Image.Resampling.LANCZOS)

            photo = ImageTk.PhotoImage(pil)

            logo_label = tb.Label(left_side, image=photo)
            logo_label.image = photo
            logo_label.pack(pady=10)

        except Exception as e:
            print(f"Не удалось загрузить иконку приложения: {e}")
            # fallback — если файл отсутствует
            tb.Label(left_side, text="🚗", font=("Segoe UI", 60)).pack(pady=10)


        tb.Label(
            left_side,
            text="Auto Tracker Pro",
            font=("Segoe UI", 28, "bold")
        ).pack(pady=5)

        tb.Label(
            left_side,
            text="Система відстеження транспортних засобів",
            font=("Segoe UI", 12)
        ).pack(pady=5)

        # Список возможностей (делаем аккуратно)
        features_frame = tb.Frame(left_side)
        features_frame.pack(pady=20)

        features = [
            "✓ Візуальні карточки авто",
            "✓ Інтерактивні карти маршрутів",
            "✓ Відстеження в реальному часі",
            "✓ Фото-галерея з каруселлю",
            "✓ Лічильник днів до прибуття",
            "✓ Аналітика та звіти"
        ]

        for feature in features:
            tb.Label(
                features_frame,
                text=feature,
                font=("Segoe UI", 11)
            ).pack(anchor="w", pady=3)

        # -------------------------------------------------------
        #                     ПРАВАЯ ЧАСТЬ (ФОРМА)
        # -------------------------------------------------------
        right_side = tb.Frame(main_container)
        right_side.pack(side="right", fill="both", expand=True)

        # Форма в виде карточки
        form_card = tb.Frame(
            right_side,
            padding=30,
            borderwidth=1,
            relief="solid"
        )
        form_card.pack(expand=True, padx=80, pady=40)

        tb.Label(
            form_card,
            text="Вхід до системи",
            font=("Segoe UI", 20, "bold")
        ).pack(pady=10)

        # ---------------- Поле логина ----------------
        tb.Label(
            form_card,
            text="Логін",
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(15, 2))

        self.entry_username = tb.Entry(form_card, width=25, font=("Segoe UI", 11))
        self.entry_username.pack(fill="x", pady=5)
        self.entry_username.bind("<Return>", lambda e: self._do_login())

        # ---------------- Поле пароля ----------------
        tb.Label(
            form_card,
            text="Пароль",
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(15, 2))

        password_frame = tb.Frame(form_card)
        password_frame.pack(fill="x", pady=5)

        self.entry_password = tb.Entry(password_frame, width=25, show="•", font=("Segoe UI", 11))
        self.entry_password.pack(side="left", fill="x", expand=True)
        self.entry_password.bind("<Return>", lambda e: self._do_login())

        # Значок показать/скрыть пароль
        self.password_visible = False

        def toggle_password():
            self.password_visible = not self.password_visible
            if self.password_visible:
                self.entry_password.config(show="")
                eye_btn.config(text="🙈")
            else:
                self.entry_password.config(show="•")
                eye_btn.config(text="👁️")

        eye_btn = tb.Button(
            password_frame,
            text="👁️",
            bootstyle="secondary-outline",
            width=3,
            command=toggle_password
        )
        eye_btn.pack(side="right", padx=5)


        # ---------------- Кнопка входа ----------------
        login_btn = tb.Button(
            form_card,
            text="Увійти",
            bootstyle="success",
            width=20,
            command=self._do_login
        )
        login_btn.pack(pady=20)

        # ---------------- Демо-кнопки ----------------
        demo_frame = tb.Frame(form_card)
        demo_frame.pack(pady=10)

        tb.Button(
            demo_frame,
            text="Демо Admin",
            bootstyle="info-outline",
            width=13,
            command=lambda: self._demo_login("admin")
        ).pack(side="left", padx=5)

        tb.Button(
            demo_frame,
            text="Демо User",
            bootstyle="secondary-outline",
            width=13,
            command=lambda: self._demo_login("user")
        ).pack(side="left", padx=5)

    def _demo_login(self, user_type):
        """Заполнение демо-данных для входа"""
        if user_type == "admin":
            self.entry_username.delete(0, 'end')
            self.entry_password.delete(0, 'end')
            self.entry_username.insert(0, "demo_user")
            self.entry_password.insert(0, "demo123")
        else:
            self.entry_username.delete(0, 'end')
            self.entry_password.delete(0, 'end')
            self.entry_username.insert(0, "user1")
            self.entry_password.insert(0, "user123")
        self.after(100, self._do_login)

    def _do_login(self):
        """Выполнение входа в систему"""
        username = self.entry_username.get().strip()
        pwd = self.entry_password.get().strip()
        
        if not username or not pwd:
            messagebox.showwarning("Помилка", "Введіть логін та пароль")
            return

        conn = safe_connect()
        if not conn:
            return
            
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT * FROM users WHERE username = %s", (username,))
            user = cur.fetchone()
            
            if not user:
                messagebox.showerror("Помилка", "Користувача не знайдено")
                cur.close()
                conn.close()
                return
                
            if not check_password_hash(user["password_hash"], pwd):
                messagebox.showerror("Помилка", "Невірний пароль")
                cur.close()
                conn.close()
                return
                
            self.current_user = user
            self.conn = conn
            
            cur.close()
            
            self._build_main_ui()
            
        except Error as e:
            messagebox.showerror("Помилка БД", f"Помилка входу: {str(e)}")
            try:
                cur.close()
                conn.close()
            except:
                pass
        finally:
            if cur:
                try:
                    cur.close()
                except:
                    pass

    def _build_main_ui(self):
        """Построение основного интерфейса"""
        for w in self.winfo_children():
            w.destroy()

        self._create_top_bar()
        self._create_main_content()

    def _create_top_bar(self):
        """Создание верхней панели"""
        topbar = tb.Frame(self, bootstyle="primary")
        topbar.pack(fill="x", padx=0, pady=0)

        left_info = tb.Frame(topbar, bootstyle="primary")
        left_info.pack(side="left", padx=12, pady=6)
        
        tb.Label(left_info, text="🚗 Auto Tracker Pro", font=("Segoe UI", 12, "bold"), 
                bootstyle="inverse-primary").pack(side="left")
        
        user_info = tb.Frame(topbar, bootstyle="primary")
        user_info.pack(side="left", padx=15, pady=6)
        
        tb.Label(user_info, text=f"Вітаємо, {self.current_user['username']}", 
                font=("Segoe UI", 10, "bold"), bootstyle="inverse-primary").pack(side="left")
        tb.Label(user_info, text=f"• Роль: {self.current_user['role']}", 
                bootstyle="inverse-primary").pack(side="left", padx=8)

        controls = tb.Frame(topbar, bootstyle="primary")
        controls.pack(side="right", padx=12, pady=6)

        tb.Button(controls, text="🌙 Тема", bootstyle="primary-outline",
                 command=self._toggle_theme).pack(side="left", padx=3)
        tb.Button(controls, text="Вийти", bootstyle="danger-outline",
                 command=self._logout).pack(side="left", padx=3)

    def _create_main_content(self):
        """Создание основного контента"""
        main_container = tb.Frame(self)
        main_container.pack(fill="both", expand=True, padx=8, pady=8)

        if self.current_user["role"] == "admin":
            self._build_admin_dashboard(main_container)
        else:
            self._build_user_dashboard(main_container)

    def _build_admin_dashboard(self, parent):
        """Построение админ-панели"""
        SIDEBAR_WIDTH = 260

        self.left_frame = tb.Frame(parent, width=SIDEBAR_WIDTH)
        self.left_frame.pack(side="left", fill="y", padx=(0, 8))
        self.left_frame.pack_propagate(False)  # фіксована ширина

        self.main_content = tb.Frame(parent)
        self.main_content.pack(side="left", fill="both", expand=True)

        self._build_admin_sidebar()
        self._show_admin_dashboard()

    def _build_admin_sidebar(self):
        """Побудова кольорової адмін-панелі"""
        
        # HEADER
        sidebar_header = tb.Frame(self.left_frame, bootstyle="primary", padding=12)
        sidebar_header.pack(fill="x", pady=(0, 12))

        tb.Label(
            sidebar_header,
            text="⚙️ Адмін-панель",
            font=("Segoe UI", 14, "bold"),
            bootstyle="inverse-primary"
        ).pack()

        # MAIN NAVIGATION
        nav_frame = tb.Frame(self.left_frame, padding=5)
        nav_frame.pack(fill="x", pady=10)

        # NEW COLORS FOR BUTTONS
        main_functions = [
            ("📊 Головна", "dashboard", "primary"),
            ("🚗 Авто", "purchases_visual", "info"),
            ("📈 Аналітика", "analytics", "warning"),
        ]

        for icon_text, table, color in main_functions:
            tb.Button(
                nav_frame,
                text=icon_text,
                bootstyle=f"{color}",
                command=lambda t=table: self._admin_navigate(t),
                padding=8
            ).pack(fill="x", pady=4)

        # SEPARATOR
        sep = ttk.Separator(nav_frame, orient='horizontal')
        sep.pack(fill='x', pady=10)

        # TABLE MANAGEMENT
        tb.Button(
            nav_frame,
            text="🛠️ Управління таблицями",
            bootstyle="danger",
            command=self._show_table_management,
            padding=8
        ).pack(fill="x", pady=4)


    def _show_table_management(self):
        """Показ диалога управления таблицами"""
        table_dialog = tb.Toplevel(self)
        table_dialog.title("Управління таблицями")
        table_dialog.geometry("400x450")
        table_dialog.transient(self)
        table_dialog.grab_set()
        
        self.center_window(table_dialog, 400, 450)

        tb.Label(table_dialog, text="⚙️ Управління таблицями", 
                font=("Segoe UI", 14, "bold")).pack(pady=15)

        tables_frame = tb.Frame(table_dialog)
        tables_frame.pack(fill="both", expand=True, padx=20, pady=10)

        tables = [
            ("👥 Користувачі", "users"),
            ("🌍 Країни", "countries"), 
            ("⚓ Порти", "ports"),
            ("🏢 Аукціони", "auctions"),
            ("📍 Локації", "locations"),
            ("🚗 Авто", "purchases"),
            ("🖼️ Фото", "purchase_images")
        ]

        for icon_text, table in tables:
            btn = tb.Button(tables_frame, text=icon_text, bootstyle="outline",
                          command=lambda t=table: [table_dialog.destroy(), self._show_table_in_main(t)])
            btn.pack(fill="x", pady=6)

        tb.Button(table_dialog, text="Закрити", bootstyle="secondary",
                 command=table_dialog.destroy).pack(pady=12)

    def _admin_navigate(self, destination):
        """Навигация по админ-панели"""
        if destination == "dashboard":
            self._show_admin_dashboard()
        elif destination == "analytics":
            self._show_analytics()
        elif destination == "purchases_visual":
            self._show_purchases_visual()
        else:
            self._show_table_in_main(destination)

    def _clear_main_content(self):
        """Очистка основного контента"""
        for w in self.main_content.winfo_children():
            w.destroy()

    def _show_admin_dashboard(self):
        """Показать админ-дашборд"""
        self._clear_main_content()
        
        welcome_card = tb.Frame(self.main_content, bootstyle="info")
        welcome_card.pack(fill="x", pady=(0,15))
        
        tb.Label(welcome_card, text="Адміністративна панель", 
                font=("Segoe UI", 18, "bold"), bootstyle="inverse-info").pack(pady=12)

        stats_frame = tb.Frame(self.main_content)
        stats_frame.pack(fill="x", pady=8)

        try:
            cur = self.conn.cursor()
            
            cur.execute("SELECT COUNT(*) FROM users")
            total_users = cur.fetchone()[0]
            
            cur.execute("SELECT COUNT(*) FROM purchases")
            total_purchases = cur.fetchone()[0]
            
            cur.execute("SELECT COUNT(*) FROM purchases WHERE status_id < 9")
            active_deliveries = cur.fetchone()[0]
            
            cur.execute("SELECT COUNT(*) FROM countries")
            total_countries = cur.fetchone()[0]
            
            cur.close()
            
            stats = [
                ("👥 Користувачі", total_users, "primary"),
                ("🚗 Всього", total_purchases, "success"),
                ("📦 Активні доставки", active_deliveries, "warning"),
                ("🌍 Країни", total_countries, "info")
            ]
            
            for text, count, style in stats:
                card = tb.Frame(stats_frame, bootstyle=style, padding=8)
                card.pack(side="left", fill="x", expand=True, padx=3)
                tb.Label(card, text=text, font=("Segoe UI", 10), 
                        bootstyle=f"inverse-{style}").pack()
                tb.Label(card, text=str(count), font=("Segoe UI", 20, "bold"),
                        bootstyle=f"inverse-{style}").pack()
                
        except Error as e:
            print(f"Помилка завантаження статистики: {e}")

        late_cars_frame = tb.LabelFrame(self.main_content, text="⚠️ Авто з запізненням", padding=12)
        late_cars_frame.pack(fill="x", pady=15)

        try:
            cur = self.conn.cursor(dictionary=True)
            cur.execute("""
                SELECT p.*, s.status_name, u.username,
                       DATEDIFF(CURDATE(), p.estimated_arrival_date) as days_late
                FROM purchases p
                LEFT JOIN statuses s ON p.status_id = s.status_id
                LEFT JOIN users u ON p.buyer_id = u.id
                WHERE p.estimated_arrival_date < CURDATE()
                AND p.is_delivered = FALSE
                AND s.status_name NOT LIKE '%Україні%'
                ORDER BY p.estimated_arrival_date ASC
            """)
            late_cars = cur.fetchall()
            cur.close()

            late_cars_scroll_container = tb.Frame(late_cars_frame)
            late_cars_scroll_container.pack(fill="both", expand=True)
            
            canvas = Canvas(late_cars_scroll_container, height=200)
            scrollbar = ttk.Scrollbar(late_cars_scroll_container, orient="vertical", command=canvas.yview)
            late_cars_content = tb.Frame(canvas)
            
            late_cars_content.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=late_cars_content, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            if late_cars:
                for car in late_cars:
                    car_frame = tb.Frame(late_cars_content)
                    car_frame.pack(fill="x", pady=6)
                    
                    tb.Label(car_frame, 
                            text=f"{car['car_make']} {car['car_model']} ({car['car_year']})",
                            font=("Segoe UI", 11, "bold")).pack(anchor="w")
                    
                    tb.Label(car_frame, 
                            text=f"Запізнення: {car['days_late']} дн. • {car['status_name']} • {car['username']}",
                            font=("Segoe UI", 9),
                            bootstyle="danger").pack(anchor="w")
                    
                    tb.Button(car_frame, text="👀 Деталі", bootstyle="outline",
                             command=lambda c=car: self._show_purchase_details(c)).pack(anchor="w", pady=3)
                    
                    ttk.Separator(car_frame, orient='horizontal').pack(fill='x', pady=3)
            else:
                tb.Label(late_cars_content, text="🎉 Немає авто з запізненням!",
                        font=("Segoe UI", 10), bootstyle="success").pack(pady=10)
                        
        except Error as e:
            tb.Label(late_cars_frame, text=f"Помилка завантаження: {e}",
                    bootstyle="danger").pack(pady=10)

    def _show_purchases_visual(self):
        """Показать покупки в виде карточек (оновлений дизайн фільтрів)"""
        self._clear_main_content()
        
        # ---------- Заголовок ----------
        header_frame = tb.Frame(self.main_content)
        header_frame.pack(fill="x", pady=(0, 8))
        
        tb.Label(
            header_frame,
            text="Куплені авто",
            font=("Segoe UI", 16, "bold")
        ).pack(side="left")

        control_frame = tb.Frame(header_frame)
        control_frame.pack(side="right")

        if self.current_user["role"] == "admin":
            tb.Button(
                control_frame,
                text="➕ Додати авто",
                bootstyle="success",
                command=self._add_new_purchase
            ).pack(side="left", padx=3)

        view_buttons = tb.Frame(control_frame)
        view_buttons.pack(side="left", padx=8)
        
        tb.Button(
            view_buttons,
            text="🎴 Картки",
            bootstyle="primary",
            command=lambda: self._show_purchases_visual()
        ).pack(side="left", padx=2)

        tb.Button(
            view_buttons,
            text="📋 Таблиця",
            bootstyle="secondary-outline",
            command=lambda: self._show_table_in_main("purchases")
        ).pack(side="left", padx=2)

        # ---------- Картка фільтрів ----------
        filters_card = tb.LabelFrame(
            self.main_content,
            text="Фільтри",
            padding=10
        )
        filters_card.pack(fill="x", pady=5)

        # Перша лінія — комбобокси
        row1 = tb.Frame(filters_card)
        row1.pack(fill="x", pady=3)

        # Статус
        status_block = tb.Frame(row1, padding=(5, 2))
        status_block.pack(side="left", padx=5, fill="x", expand=True)

        tb.Label(
            status_block,
            text="⚙️ Статус",
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")

        self.status_filter = tb.StringVar(value="all")
        status_combo = tb.Combobox(
            status_block,
            textvariable=self.status_filter,
            values=["all", "bought_auction", "paid", "to_port", "at_port",
                    "in_sea", "in_klaipeda", "to_ukraine", "cleared_customs", "in_ukraine"],
            state="readonly",
            width=18
        )
        status_combo.pack(fill="x", pady=2)
        status_combo.bind("<<ComboboxSelected>>", self._load_purchases_cards)

        # Країна
        country_block = tb.Frame(row1, padding=(5, 2))
        country_block.pack(side="left", padx=5, fill="x", expand=True)

        tb.Label(
            country_block,
            text="🌍 Країна",
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")

        self.country_filter = tb.StringVar(value="all")
        try:
            cur = self.conn.cursor()
            cur.execute("SELECT country_name FROM countries ORDER BY country_name")
            countries = [row[0] for row in cur.fetchall()]
            cur.close()

            country_combo = tb.Combobox(
                country_block,
                textvariable=self.country_filter,
                values=["all"] + countries,
                state="readonly",
                width=18
            )
            country_combo.pack(fill="x", pady=2)
            country_combo.bind("<<ComboboxSelected>>", self._load_purchases_cards)
        except Error as e:
            print(f"Помилка завантаження країн: {e}")

        # Рік
        year_block = tb.Frame(row1, padding=(5, 2))
        year_block.pack(side="left", padx=5, fill="x")

        tb.Label(
            year_block,
            text="📅 Рік",
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")

        self.year_filter = tb.StringVar(value="all")
        current_year = datetime.datetime.now().year
        years = [str(y) for y in range(current_year - 10, current_year + 1)]
        year_combo = tb.Combobox(
            year_block,
            textvariable=self.year_filter,
            values=["all"] + years,
            state="readonly",
            width=10
        )
        year_combo.pack(fill="x", pady=2)
        year_combo.bind("<<ComboboxSelected>>", self._load_purchases_cards)

        # Друга лінія — пошук + кнопки
        row2 = tb.Frame(filters_card)
        row2.pack(fill="x", pady=5)

        self.search_var = tb.StringVar()
        search_entry = tb.Entry(
            row2,
            textvariable=self.search_var,
            width=40,
            font=("Segoe UI", 10)
        )
        search_entry.pack(side="left", padx=5, fill="x", expand=True)
        search_entry.insert(0, "Пошук по VIN, марці, моделі...")

        def on_search_focus_in(e):
            if search_entry.get() == "Пошук по VIN, марці, моделі...":
                search_entry.delete(0, 'end')
        
        def on_search_focus_out(e):
            if not search_entry.get():
                search_entry.insert(0, "Пошук по VIN, марці, моделі...")

        search_entry.bind("<FocusIn>", on_search_focus_in)
        search_entry.bind("<FocusOut>", on_search_focus_out)
        search_entry.bind("<KeyRelease>", lambda e: self._filter_purchases_cards())

        tb.Button(
            row2,
            text="🔍 Пошук",
            bootstyle="info",
            command=self._filter_purchases_cards
        ).pack(side="left", padx=5)

        tb.Button(
            row2,
            text="🗑️ Скинути",
            bootstyle="danger-outline",
            command=self._reset_filters
        ).pack(side="right", padx=5)

        # ---------- Контейнер карток ----------
        cards_container = tb.Frame(self.main_content)
        cards_container.pack(fill="both", expand=True, pady=8)
        
        canvas = Canvas(cards_container)
        scrollbar = ttk.Scrollbar(cards_container, orient="vertical", command=canvas.yview)
        self.cards_frame = tb.Frame(canvas)
        
        self.cards_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.cards_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self._load_purchases_cards()

    
    def _reset_filters(self):
        """Сброс фильтров"""
        self.status_filter.set("all")
        self.country_filter.set("all")
        self.year_filter.set("all")
        self.search_var.set("")
        self._load_purchases_cards()
    
    def _load_purchases_cards(self, event=None):
        """Загрузка карточек покупок"""
        for w in self.cards_frame.winfo_children():
            w.destroy()
        
        try:
            cur = self.conn.cursor(dictionary=True)
            
            query = """
                SELECT p.*, c.country_name, a.auction_name, l.location_name, 
                       s.status_name, u.username, port.port_name,
                       EXISTS(SELECT 1 FROM purchase_images WHERE purchase_id = p.purchase_id) as has_images,
                       (SELECT image_url FROM purchase_images WHERE purchase_id = p.purchase_id LIMIT 1) as first_image_path
                FROM purchases p
                LEFT JOIN countries c ON p.country_id = c.country_id
                LEFT JOIN auctions a ON p.auction_id = a.auction_id
                LEFT JOIN locations l ON p.location_id = l.location_id
                LEFT JOIN statuses s ON p.status_id = s.status_id
                LEFT JOIN users u ON p.buyer_id = u.id
                LEFT JOIN locations loc ON p.location_id = loc.location_id
                LEFT JOIN ports port ON loc.default_port_id = port.port_id
                WHERE 1=1
            """
            
            params = []
            
            status_filter = self.status_filter.get()
            if status_filter != "all":
                query += " AND s.status_key = %s"
                params.append(status_filter)
            
            country_filter = self.country_filter.get()
            if country_filter != "all":
                query += " AND c.country_name = %s"
                params.append(country_filter)
            
            year_filter = self.year_filter.get()
            if year_filter != "all":
                query += " AND p.car_year = %s"
                params.append(int(year_filter))
            
            query += " ORDER BY p.purchase_date DESC"
            
            cur.execute(query, params)
            purchases = cur.fetchall()
            cur.close()
            
            row_frame = None
            for idx, purchase in enumerate(purchases):
                if idx % 3 == 0:
                    row_frame = tb.Frame(self.cards_frame)
                    row_frame.pack(fill="x", pady=3)
                
                card = CarCard(row_frame, purchase, on_click=self._show_purchase_details)
                card.pack(side="left", padx=8, fill="both", expand=True)
            
            if not purchases:
                tb.Label(self.cards_frame, text="Немає покупок за обраними фільтрами", 
                        font=("Segoe UI", 12)).pack(pady=40)
                
        except Error as e:
            messagebox.showerror("Помилка", f"Помилка завантаження: {e}")
    
    def _filter_purchases_cards(self):
        """Фильтрация карточек покупок"""
        search_text = self.search_var.get().lower()
        if search_text == "Пошук по VIN, марці, моделі...":
            search_text = ""
        
        for w in self.cards_frame.winfo_children():
            w.destroy()
        
        try:
            cur = self.conn.cursor(dictionary=True)
            
            query = """
                SELECT p.*, c.country_name, a.auction_name, l.location_name, 
                       s.status_name, u.username, port.port_name,
                       EXISTS(SELECT 1 FROM purchase_images WHERE purchase_id = p.purchase_id) as has_images,
                       (SELECT image_url FROM purchase_images WHERE purchase_id = p.purchase_id LIMIT 1) as first_image_path
                FROM purchases p
                LEFT JOIN countries c ON p.country_id = c.country_id
                LEFT JOIN auctions a ON p.auction_id = a.auction_id
                LEFT JOIN locations l ON p.location_id = l.location_id
                LEFT JOIN statuses s ON p.status_id = s.status_id
                LEFT JOIN users u ON p.buyer_id = u.id
                LEFT JOIN locations loc ON p.location_id = loc.location_id
                LEFT JOIN ports port ON loc.default_port_id = port.port_id
                WHERE 1=1
            """
            
            params = []
            
            status_filter = self.status_filter.get()
            if status_filter != "all":
                query += " AND s.status_key = %s"
                params.append(status_filter)
            
            country_filter = self.country_filter.get()
            if country_filter != "all":
                query += " AND c.country_name = %s"
                params.append(country_filter)
            
            year_filter = self.year_filter.get()
            if year_filter != "all":
                query += " AND p.car_year = %s"
                params.append(int(year_filter))
            
            if search_text:
                query += " AND (LOWER(p.vin_number) LIKE %s OR LOWER(p.car_make) LIKE %s OR LOWER(p.car_model) LIKE %s)"
                params.extend([f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"])
            
            query += " ORDER BY p.purchase_date DESC"
            
            cur.execute(query, params)
            purchases = cur.fetchall()
            cur.close()
            
            row_frame = None
            for idx, purchase in enumerate(purchases):
                if idx % 3 == 0:
                    row_frame = tb.Frame(self.cards_frame)
                    row_frame.pack(fill="x", pady=3)
                
                card = CarCard(row_frame, purchase, on_click=self._show_purchase_details)
                card.pack(side="left", padx=8, fill="both", expand=True)
            
            if not purchases:
                tb.Label(self.cards_frame, text="Нічого не знайдено", 
                        font=("Segoe UI", 12)).pack(pady=40)
                
        except Error as e:
            messagebox.showerror("Помилка", f"Помилка пошуку: {e}")

    def _show_purchase_details(self, purchase):
        """Показать детали покупки"""
        self.selected_purchase = purchase
        
        details_window = tb.Toplevel(self)
        details_window.title(f"Деталі: {purchase['car_make']} {purchase['car_model']}")
        details_window.geometry("1300x700")
        details_window.transient(self)
        
        self.center_window(details_window, 1300, 700)
        
        header = tb.Frame(details_window, bootstyle="primary", padding=12)
        header.pack(fill="x")
        
        tb.Label(header, 
                text=f"🚗 {purchase['car_make']} {purchase['car_model']} ({purchase['car_year']})",
                font=("Segoe UI", 16, "bold"),
                bootstyle="inverse-primary").pack(side="left")
        
        if self.current_user["role"] == "admin":
            admin_buttons = tb.Frame(header)
            admin_buttons.pack(side="right")
            
            tb.Button(admin_buttons, text="✏️ Редагувати", bootstyle="warning",
                    command=lambda: self._edit_purchase(purchase, details_window)).pack(side="left", padx=2)
            tb.Button(admin_buttons, text="🔄 Змінити статус", bootstyle="info",
                    command=lambda: self._quick_status_change(purchase)).pack(side="left", padx=2)
        
        tb.Button(header, text="❌ Закрити", bootstyle="secondary",
                command=details_window.destroy).pack(side="right", padx=3)
        
        content = tb.Frame(details_window, padding=15)
        content.pack(fill="both", expand=True)
        
        left_panel = tb.Frame(content, bootstyle="light", padding=12)
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 8))
        
        days_counter = DaysCounterWidget(left_panel, purchase)
        days_counter.pack(fill="x", pady=(0, 8))
        
        info_section = tb.LabelFrame(left_panel, text="Інформація про авто", 
                                    padding=8)
        info_section.pack(fill="x", pady=3)
        
        info_items = [
            ("VIN:", purchase.get('vin_number', 'N/A')),
            ("Марка:", purchase.get('car_make', 'N/A')),
            ("Модель:", purchase.get('car_model', 'N/A')),
            ("Рік:", purchase.get('car_year', 'N/A')),
            ("Ціна:", f"${purchase.get('price_usd', 0):,.2f}"),
            ("Покупець:", purchase.get('username', 'N/A')),
        ]
        
        for label, value in info_items:
            row = tb.Frame(info_section)
            row.pack(fill="x", pady=2)
            tb.Label(row, text=label, font=("Segoe UI", 9, "bold"), 
                    width=10, anchor="w").pack(side="left")
            tb.Label(row, text=str(value), font=("Segoe UI", 9)).pack(side="left")
        
        delivery_section = tb.LabelFrame(left_panel, text="Доставка", 
                                        padding=8)
        delivery_section.pack(fill="x", pady=3)
        
        delivery_items = [
            ("Статус:", purchase.get('status_name', 'N/A')),
            ("Країна:", purchase.get('country_name', 'N/A')),
            ("Аукціон:", purchase.get('auction_name', 'N/A')),
            ("Локація:", purchase.get('location_name', 'N/A')),
            ("Порт:", purchase.get('port_name', 'N/A')),
            ("Дата покупки:", purchase.get('purchase_date', 'N/A')),
            ("Очікувана дата:", purchase.get('estimated_arrival_date', 'N/A')),
        ]
        
        for label, value in delivery_items:
            row = tb.Frame(delivery_section)
            row.pack(fill="x", pady=2)
            tb.Label(row, text=label, font=("Segoe UI", 9, "bold"), 
                    width=12, anchor="w").pack(side="left")
            
            if label == "Статус:":
                status_color = "success" if "україні" in str(value).lower() else "warning"
                tb.Label(row, text=str(value), font=("Segoe UI", 9, "bold"),
                        bootstyle=status_color).pack(side="left")
            else:
                tb.Label(row, text=str(value), font=("Segoe UI", 9)).pack(side="left")
        
        if purchase.get('notes'):
            notes_section = tb.LabelFrame(left_panel, text="Примітки", 
                                         padding=8)
            notes_section.pack(fill="x", pady=3)
            
            notes_text = tb.Text(notes_section, height=3, font=("Segoe UI", 8))
            notes_text.insert("1.0", purchase['notes'])
            notes_text.config(state="disabled")
            notes_text.pack(fill="x")
        
        right_panel = tb.Frame(content)
        right_panel.pack(side="right", fill="both", expand=True)
        
        view_switch_frame = tb.Frame(right_panel)
        view_switch_frame.pack(fill="x", pady=3)
        
        self.details_view_mode = tb.StringVar(value="photos")
        
        tb.Radiobutton(view_switch_frame, text="🖼️ Фото", variable=self.details_view_mode,
                      value="photos", command=self._update_details_view).pack(side="left", padx=3)
        tb.Radiobutton(view_switch_frame, text="🗺️ Карта", variable=self.details_view_mode,
                      value="map", command=self._update_details_view).pack(side="left", padx=3)
        
        self.details_view_container = tb.Frame(right_panel)
        self.details_view_container.pack(fill="both", expand=True)
        
        self._update_details_view()
    
    def _update_details_view(self):
        """Обновление вида деталей (фото/карта)"""
        for w in self.details_view_container.winfo_children():
            w.destroy()
        
        if self.details_view_mode.get() == "photos":
            carousel = ImageCarousel(self.details_view_container, 
                                   self.selected_purchase['purchase_id'], 
                                   self.conn, self.current_user)
            carousel.pack(fill="both", expand=True)
        else:
            map_widget = MapWidget(self.details_view_container, self.selected_purchase)
            map_widget.pack(fill="both", expand=True)

    def _edit_purchase(self, purchase, parent_window):
        """Редактирование покупки"""
        def save_data(updated_data, mode):
            try:
                cur = self.conn.cursor()
                
                if not updated_data:
                    messagebox.showwarning("Помилка", "Немає даних для оновлення")
                    return
                
                set_clause = ', '.join([f"`{k}`=%s" for k in updated_data.keys()])
                sql = f"UPDATE purchases SET {set_clause} WHERE purchase_id=%s"
                
                values_list = list(updated_data.values())
                values_list.append(purchase['purchase_id'])
                
                cur.execute(sql, tuple(values_list))
                self.conn.commit()
                cur.close()
                
                parent_window.destroy()
                self._show_purchases_visual()
                
            except Error as e:
                messagebox.showerror("Помилка", f"Помилка оновлення: {str(e)}")

        ModernCRUDDialog(self, f"Редагувати: {purchase['car_make']} {purchase['car_model']}", 
                        "purchases", "edit", purchase, on_save=save_data)

    def _quick_status_change(self, purchase):
        """Быстрое изменение статуса покупки (оновлений дизайн)"""
        status_dialog = tb.Toplevel(self)
        status_dialog.title("Швидка зміна статусу")
        status_dialog.geometry("380x380")
        status_dialog.transient(self)
        status_dialog.grab_set()
        
        self.center_window(status_dialog, 380, 380)
        
        # Заголовок
        header = tb.Frame(status_dialog, bootstyle="primary", padding=10)
        header.pack(fill="x")
        
        tb.Label(
            header,
            text="⚙️ Швидка зміна статусу",
            font=("Segoe UI", 12, "bold"),
            bootstyle="inverse-primary"
        ).pack(side="left")
        
        # Поточний статус
        current_status_frame = tb.Frame(status_dialog, padding=10)
        current_status_frame.pack(fill="x")
        
        tb.Label(
            current_status_frame,
            text="Поточний статус:",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w")
        
        tb.Label(
            current_status_frame,
            text=purchase.get('status_name', 'Невідомо'),
            font=("Segoe UI", 10),
            bootstyle="success"
        ).pack(anchor="w", pady=3)

        # Карточка зі списком статусів
        list_card = tb.LabelFrame(
            status_dialog,
            text="Оберіть новий статус",
            padding=10
        )
        list_card.pack(fill="both", expand=True, padx=10, pady=5)
        
        try:
            cur = self.conn.cursor(dictionary=True)
            cur.execute("SELECT status_id, status_name FROM statuses ORDER BY order_index")
            statuses = cur.fetchall()
            cur.close()
            
            status_var = tb.StringVar(value=purchase['status_name'])

            # Список статусів
            for status in statuses:
                row = tb.Frame(list_card, padding=3)
                row.pack(fill="x", pady=1)

                rb = tb.Radiobutton(
                    row,
                    text=status['status_name'],
                    variable=status_var,
                    value=status['status_name']
                )
                rb.pack(side="left", anchor="w")

                # Позначаємо поточний статус
                if status['status_name'] == purchase['status_name']:
                    tb.Label(
                        row,
                        text="Поточний",
                        font=("Segoe UI", 8, "bold"),
                        bootstyle="success"
                    ).pack(side="right")

            def save_status():
                try:
                    selected_status = None
                    for status in statuses:
                        if status['status_name'] == status_var.get():
                            selected_status = status
                            break
                    
                    if selected_status:
                        cur = self.conn.cursor()
                        cur.execute("""
                            UPDATE purchases 
                            SET status_id = %s 
                            WHERE purchase_id = %s
                        """, (selected_status['status_id'], purchase['purchase_id']))
                        self.conn.commit()
                        cur.close()
                        
                        messagebox.showinfo("Успіх", "Статус успішно оновлено!")
                        status_dialog.destroy()
                        self._show_purchases_visual()
                        
                except Error as e:
                    messagebox.showerror("Помилка", f"Помилка оновлення статусу: {str(e)}")
            
            # Кнопки знизу
            btn_frame = tb.Frame(status_dialog, padding=10)
            btn_frame.pack(fill="x")
            
            tb.Button(
                btn_frame,
                text="💾 Зберегти",
                bootstyle="success",
                command=save_status
            ).pack(side="left", padx=3)

            tb.Button(
                btn_frame,
                text="❌ Скасувати",
                bootstyle="secondary",
                command=status_dialog.destroy
            ).pack(side="right", padx=3)
            
        except Error as e:
            messagebox.showerror("Помилка", f"Помилка завантаження статусів: {e}")

    
    def _add_new_purchase(self):
        """Добавление новой покупки"""
        def save_data(data, mode):
            try:
                cur = self.conn.cursor()
                
                if not data:
                    messagebox.showwarning("Помилка", "Немає даних для збереження")
                    return
                
                columns = ', '.join([f"`{k}`" for k in data.keys()])
                placeholders = ', '.join(['%s'] * len(data))
                sql = f"INSERT INTO purchases ({columns}) VALUES ({placeholders})"
                
                cur.execute(sql, tuple(data.values()))
                self.conn.commit()
                cur.close()
                self._show_purchases_visual()
            except Error as e:
                messagebox.showerror("Помилка", f"Помилка додавання: {str(e)}")

        ModernCRUDDialog(self, "Додати нову покупку", "purchases", "add", on_save=save_data)

    def _show_table_in_main(self, table):
        """Показать таблицу в основном окне"""
        self._clear_main_content()
        self.current_table = table
        
        table_frame = tb.Frame(self.main_content)
        table_frame.pack(fill="both", expand=True, padx=8, pady=8)
        
        tb.Label(table_frame, text=f"Таблиця: {table}", 
                font=("Segoe UI", 16, "bold")).pack(pady=(0, 8))

        if table == "purchases":
            view_frame = tb.Frame(table_frame)
            view_frame.pack(fill="x", pady=3)
            
            tb.Button(view_frame, text="🎴 Переглянути картки", bootstyle="info",
                     command=lambda: self._show_purchases_visual()).pack(side="left", padx=3)

        tree_container = tb.Frame(table_frame)
        tree_container.pack(fill="both", expand=True, pady=3)

        try:
            if table == "purchases":
                query = """
                    SELECT p.*, 
                           c.country_name, 
                           a.auction_name, 
                           l.location_name, 
                           s.status_name, 
                           u.username,
                           EXISTS(SELECT 1 FROM purchase_images WHERE purchase_id = p.purchase_id) as has_images
                    FROM purchases p
                    LEFT JOIN countries c ON p.country_id = c.country_id
                    LEFT JOIN auctions a ON p.auction_id = a.auction_id
                    LEFT JOIN locations l ON p.location_id = l.location_id
                    LEFT JOIN statuses s ON p.status_id = s.status_id
                    LEFT JOIN users u ON p.buyer_id = u.id
                    ORDER BY p.purchase_date DESC
                    LIMIT 100
                """
            else:
                query = f"SELECT * FROM `{table}` LIMIT 100"
                
            cur = self.conn.cursor()
            cur.execute(query)
            rows = cur.fetchall()
            cols = [description[0] for description in cur.description]
            cur.close()
        except Error as e:
            tb.Label(tree_container, text=f"Помилка завантаження таблиці: {e}").pack()
            return

        style = ttk.Style()
        style.configure("Treeview", 
                       background=self.style.colors.bg if hasattr(self.style.colors, 'bg') else "#ffffff",
                       foreground=self.style.colors.fg if hasattr(self.style.colors, 'fg') else "#000000",
                       fieldbackground=self.style.colors.bg if hasattr(self.style.colors, 'bg') else "#ffffff")
        
        style.configure("Treeview.Heading",
                       background=self.style.colors.primary if hasattr(self.style.colors, 'primary') else "#007bff",
                       foreground="white",
                       relief="flat")
        
        tree = ttk.Treeview(tree_container, columns=cols, show="headings", height=15, style="Treeview")
        
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        for col in cols:
            tree.heading(col, text=col)
            if col == 'has_images':
                tree.column(col, width=70, minwidth=50, anchor="center")
            else:
                tree.column(col, width=100, minwidth=70, anchor="w")

        for i, row in enumerate(rows):
            values = list(row)
            if 'has_images' in cols:
                has_images_index = cols.index('has_images')
                if values[has_images_index]:
                    values[has_images_index] = "📷"
                else:
                    values[has_images_index] = ""
            
            tree.insert("", "end", values=values, tags=('even',) if i % 2 == 0 else ('odd',))

        tree.tag_configure('odd', 
                          background=self.style.colors.light if hasattr(self.style.colors, 'light') else '#f8f9fa',
                          foreground=self.style.colors.fg if hasattr(self.style.colors, 'fg') else '#000000')
        tree.tag_configure('even', 
                          background=self.style.colors.bg if hasattr(self.style.colors, 'bg') else '#ffffff',
                          foreground=self.style.colors.fg if hasattr(self.style.colors, 'fg') else '#000000')

        btn_frame = tb.Frame(table_frame)
        btn_frame.pack(fill="x", pady=8)

        def refresh_table():
            try:
                tree.delete(*tree.get_children())
                
                if table == "purchases":
                    query = """
                        SELECT p.*, 
                               c.country_name, 
                               a.auction_name, 
                               l.location_name, 
                               s.status_name, 
                               u.username,
                               EXISTS(SELECT 1 FROM purchase_images WHERE purchase_id = p.purchase_id) as has_images
                        FROM purchases p
                        LEFT JOIN countries c ON p.country_id = c.country_id
                        LEFT JOIN auctions a ON p.auction_id = a.auction_id
                        LEFT JOIN locations l ON p.location_id = l.location_id
                        LEFT JOIN statuses s ON p.status_id = s.status_id
                        LEFT JOIN users u ON p.buyer_id = u.id
                        ORDER BY p.purchase_date DESC
                        LIMIT 100
                    """
                else:
                    query = f"SELECT * FROM `{table}` LIMIT 100"
                    
                cur = self.conn.cursor()
                cur.execute(query)
                rows = cur.fetchall()
                
                for i, row in enumerate(rows):
                    values = list(row)
                    if 'has_images' in cols:
                        has_images_index = cols.index('has_images')
                        if values[has_images_index]:
                            values[has_images_index] = "📷"
                        else:
                            values[has_images_index] = ""
                    
                    tree.insert("", "end", values=values, tags=('even',) if i % 2 == 0 else ('odd',))
                cur.close()
                messagebox.showinfo("Оновлено", "Таблицю оновлено")
            except Error as e:
                messagebox.showerror("Помилка", f"Помилка оновлення: {str(e)}")

        def add_record():
            def save_data(data, mode):
                try:
                    cur = self.conn.cursor()
                    
                    if not data:
                        messagebox.showwarning("Помилка", "Немає даних для збереження")
                        return
                    
                    columns = ', '.join([f"`{k}`" for k in data.keys()])
                    placeholders = ', '.join(['%s'] * len(data))
                    sql = f"INSERT INTO `{table}` ({columns}) VALUES ({placeholders})"
                    
                    cur.execute(sql, tuple(data.values()))
                    self.conn.commit()
                    cur.close()
                    refresh_table()
                except Error as e:
                    messagebox.showerror("Помилка", f"Помилка додавання: {str(e)}")

            ModernCRUDDialog(self, f"Додати запис у {table}", table, "add", on_save=save_data)

        def edit_record():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Помилка", "Оберіть рядок для редагування")
                return
            
            pk = self._get_primary_key(table)
            if not pk:
                messagebox.showerror("Помилка", "Не знайдено первинний ключ")
                return
            
            item = tree.item(selected[0])
            values = item["values"]
            
            data = {}
            for i, col in enumerate(cols):
                if col == 'has_images':
                    continue
                data[col] = values[i] if i < len(values) else ""

            def save_data(updated_data, mode):
                try:
                    cur = self.conn.cursor()
                    
                    if not updated_data:
                        messagebox.showwarning("Помилка", "Немає даних для оновлення")
                        return
                    
                    set_clause = ', '.join([f"`{k}`=%s" for k in updated_data.keys()])
                    sql = f"UPDATE `{table}` SET {set_clause} WHERE `{pk}`=%s"
                    
                    values_list = list(updated_data.values())
                    values_list.append(data[pk])
                    
                    cur.execute(sql, tuple(values_list))
                    self.conn.commit()
                    cur.close()
                    refresh_table()
                except Error as e:
                    messagebox.showerror("Помилка", f"Помилка оновлення: {str(e)}")

            ModernCRUDDialog(self, f"Редагувати запис у {table}", table, "edit", data, on_save=save_data)

        def delete_record():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Помилка", "Оберіть рядки для видалення")
                return
            
            pk = self._get_primary_key(table)
            if not pk:
                messagebox.showerror("Помилка", "Не знайдено первинний ключ")
                return
            
            if not messagebox.askyesno("Підтвердження", 
                                    f"Видалити {len(selected)} обраних записів?"):
                return
            
            try:
                cur = self.conn.cursor()
                success_count = 0
                
                for item in selected:
                    vals = tree.item(item)["values"]
                    pk_index = cols.index(pk)
                    pk_value = vals[pk_index]
                    
                    try:
                        cur.execute(f"DELETE FROM `{table}` WHERE `{pk}`=%s", (pk_value,))
                        success_count += 1
                    except Error as e:
                        print(f"Помилка видалення запису {pk_value}: {e}")
                        continue
                
                self.conn.commit()
                cur.close()
                
                if success_count > 0:
                    messagebox.showinfo("Успіх", f"Видалено {success_count} записів!")
                    refresh_table()
                    
            except Error as e:
                messagebox.showerror("Помилка", f"Помилка видалення: {str(e)}")

        if self.current_user["role"] == "admin":
            tb.Button(btn_frame, text="🔄 Оновити", bootstyle="info",
                    command=refresh_table).pack(side="left", padx=3)
            tb.Button(btn_frame, text="➕ Додати", bootstyle="success",
                    command=add_record).pack(side="left", padx=3)
            tb.Button(btn_frame, text="✏️ Редагувати", bootstyle="warning",
                    command=edit_record).pack(side="left", padx=3)
            tb.Button(btn_frame, text="🗑️ Видалити", bootstyle="danger",
                    command=delete_record).pack(side="left", padx=3)
            
            export_frame = tb.Frame(btn_frame)
            export_frame.pack(side="right", padx=3)
            
            tb.Button(export_frame, text="📊 Excel", bootstyle="success-outline",
                     command=lambda: self._export_current_table('excel')).pack(side="left", padx=2)
            tb.Button(export_frame, text="📄 CSV", bootstyle="warning-outline",
                     command=lambda: self._export_current_table('csv')).pack(side="left", padx=2)
        
        tb.Button(btn_frame, text="← Назад", bootstyle="dark",
                command=self._show_admin_dashboard if self.current_user["role"] == "admin" else self._show_user_dashboard).pack(side="right", padx=3)

    def _show_analytics(self):
        """Показать аналитику (улучшенная версия)"""
        self._clear_main_content()

        analytics_frame = tb.Frame(self.main_content)
        analytics_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # ---------------- HEADER ----------------
        header_frame = tb.Frame(analytics_frame)
        header_frame.pack(fill="x", pady=(0, 10))

        tb.Label(
            header_frame,
            text="📊 Аналітика та звіти",
            font=("Segoe UI", 18, "bold")
        ).pack(side="left")

        # кнопка создать отчёт
        tb.Button(
            header_frame,
            text="📄 Сформувати звіт",
            bootstyle="info",
            padding=5,
            command=self._open_report_range_dialog
        ).pack(side="right")


        # ---------------- QUERY ----------------
        try:
            cur = self.conn.cursor()
            cur.execute("""
                SELECT s.status_name, COUNT(p.purchase_id) as count
                FROM statuses s
                LEFT JOIN purchases p ON s.status_id = p.status_id
                GROUP BY s.status_id, s.status_name
                ORDER BY s.order_index
            """)
            status_stats = cur.fetchall()
            cur.close()

            max_value = max((c for (_, c) in status_stats), default=1)

            stats_frame = tb.Frame(analytics_frame)
            stats_frame.pack(fill="x", pady=10)

            LABEL_WIDTH = 28
            NUMBER_WIDTH = 4

            for status_name, count in status_stats:
                row = tb.Frame(stats_frame)
                row.pack(fill="x", pady=4)

                name_lower = status_name.lower()
                if "куп" in name_lower:
                    style = "info"
                elif "порт" in name_lower:
                    style = "warning"
                elif "мор" in name_lower:
                    style = "primary"
                elif "укра" in name_lower:
                    style = "success"
                else:
                    style = "secondary"

                # статус
                tb.Label(
                    row,
                    text=status_name,
                    width=LABEL_WIDTH,
                    anchor="w",
                    font=("Segoe UI", 11)
                ).pack(side="left")

                # правильный процент
                percent = int((count / max_value) * 100) if max_value else 0

                pb = tb.Progressbar(
                    row,
                    value=percent,
                    maximum=100,
                    bootstyle=style
                )
                pb.pack(side="left", fill="x", expand=True, padx=10, ipady=2)

                tb.Label(
                    row,
                    text=str(count),
                    width=NUMBER_WIDTH,
                    anchor="e",
                    font=("Segoe UI", 11, "bold")
                ).pack(side="right")

        except Error as e:
            tb.Label(analytics_frame, text=f"Помилка завантаження аналітики: {e}",
                    bootstyle="danger").pack()

        # BACK
        tb.Button(
            analytics_frame,
            text="← Назад",
            bootstyle="dark",
            padding=5,
            command=self._show_admin_dashboard if self.current_user["role"] == "admin" else self._show_user_dashboard
        ).pack(side="bottom", pady=10)


    def _open_report_range_dialog(self):
        dlg = tb.Toplevel(self)
        dlg.title("Створити звіт за період")
        dlg.geometry("420x300")
        dlg.transient(self)
        dlg.grab_set()
        self.center_window(dlg, 420, 370)

        selected = {"from": None, "to": None, "file": None}

        # ---------- CARD ----------
        card = tb.Frame(dlg, padding=15, borderwidth=1, relief="solid")
        card.pack(fill="both", expand=True, padx=15, pady=15)

        tb.Label(card, text="📅 Оберіть діапазон дат", font=("Segoe UI", 14, "bold")).pack(pady=(0, 10))

        # ---------- FROM DATE ----------
        frm_section = tb.Frame(card)
        frm_section.pack(fill="x", pady=6)

        tb.Label(frm_section, text="З дати:", font=("Segoe UI", 11)).pack(anchor="w")

        from_btn = tb.Button(frm_section, text="Не обрано",
                            bootstyle="secondary-outline", width=26)
        from_btn.pack(pady=3)

        def pick_from():
            cal = CalendarDialog(self)
            self.wait_window(cal.dialog)
            if cal.result:
                selected["from"] = cal.result
                from_btn.configure(text=str(cal.result))

        from_btn.configure(command=pick_from)

        # ---------- TO DATE ----------
        to_section = tb.Frame(card)
        to_section.pack(fill="x", pady=6)

        tb.Label(to_section, text="По дату:", font=("Segoe UI", 11)).pack(anchor="w")

        to_btn = tb.Button(to_section, text="Не обрано",
                        bootstyle="secondary-outline", width=26)
        to_btn.pack(pady=3)

        def pick_to():
            cal = CalendarDialog(self)
            self.wait_window(cal.dialog)
            if cal.result:
                selected["to"] = cal.result
                to_btn.configure(text=str(cal.result))

        to_btn.configure(command=pick_to)

        # ---- FILE SAVE PATH (ONLY ONE CHOICE) ----
        tb.Label(card, text="Файл звіту:", font=("Segoe UI", 11)).pack(pady=(10, 3))

        file_btn = tb.Button(card, text="📁 Обрати місце збереження", bootstyle="info")
        file_btn.pack()

        def pick_file():
            file = filedialog.asksaveasfilename(
                title="Зберегти звіт",
                defaultextension=".csv",
                filetypes=[
                    ("CSV файл", "*.csv"),
                    ("Excel файл", "*.xlsx"),
                    ("Всі файли", "*.*")
                ]
            )
            if file:
                selected["file"] = file
                file_btn.configure(text=file)

        file_btn.configure(command=pick_file)

        # ---- GENERATE ----
        def generate():
            if not selected["from"] or not selected["to"]:
                messagebox.showwarning("Помилка", "Оберіть обидві дати.")
                return

            if selected["from"] > selected["to"]:
                messagebox.showwarning("Помилка", "Дата 'З' не може бути більшою за 'По'.")
                return

            if not selected["file"]:
                messagebox.showwarning("Помилка", "Оберіть шлях збереження звіту.")
                return

            dlg.destroy()

            # Передаём параметры
            self._generate_report_range({
                "from": selected["from"],
                "to": selected["to"],
                "file": selected["file"]
            })

            self.report_range = {"from": str(selected["from"]), "to": str(selected["to"])}
            self._show_analytics()

        tb.Button(
            card,
            text="📄 Сформувати звіт",
            bootstyle="success",
            width=20,
            command=generate
        ).pack(pady=20)


    def _generate_report_range(self, data):
        date_from = str(data["from"])
        date_to = str(data["to"])
        file = data["file"]

        ext = os.path.splitext(file)[1].lower()

        try:
            # ----- 1) Получаем все колонки таблицы purchases -----
            cur = self.conn.cursor()  # обычный cursor!
            cur.execute("SHOW COLUMNS FROM purchases")
            columns = [col[0] for col in cur.fetchall()]
            cur.close()

            # ----- 2) Забираем данные за период -----
            query = f"""
                SELECT {", ".join(columns)}
                FROM purchases
                WHERE DATE(purchase_date) BETWEEN %s AND %s
                ORDER BY purchase_date ASC
            """

            cur = self.conn.cursor(dictionary=True)
            cur.execute(query, (date_from, date_to))
            rows = cur.fetchall()
            cur.close()

            total = len(rows)

            if not rows:
                messagebox.showinfo("Звіт", "Немає покупок за обраний період.")
                return

            # ----- 3) Сохранение CSV -----
            if ext == ".csv":
                import csv
                with open(file, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=columns)
                    writer.writeheader()
                    writer.writerows(rows)

                    f.write("\n")
                    f.write(
                        f"За період з {date_from} по {date_to} було куплено {total} автомобілів.\n"
                    )

            # ----- 4) Сохранение Excel -----
            elif ext == ".xlsx":
                import pandas as pd
                from openpyxl import load_workbook

                df_pd = pd.DataFrame(rows)
                df_pd.to_excel(file, index=False)

                wb = load_workbook(file)
                ws = wb.active
                ws.append([])
                ws.append([f"За період з {date_from} по {date_to} було куплено {total} автомобілів."])
                wb.save(file)

            messagebox.showinfo("Готово", f"Звіт збережено:\n{file}")

        except Exception as e:
            import traceback
            messagebox.showerror("Помилка", traceback.format_exc())



    def _build_user_dashboard(self, parent):
        """Построение пользовательской панели"""
        SIDEBAR_WIDTH = 260

        self.left_frame = tb.Frame(parent, width=SIDEBAR_WIDTH)
        self.left_frame.pack(side="left", fill="y", padx=(0, 8))
        self.left_frame.pack_propagate(False)  # фіксована ширина

        self.main_content = tb.Frame(parent)
        self.main_content.pack(side="left", fill="both", expand=True)

        self._build_user_sidebar()
        self._show_user_dashboard()

    def _build_user_sidebar(self):
        """Побудова оновленої бокової панелі користувача (стиль як у адмін-панелі)"""

        # HEADER — делаем аналогично admin sidebar
        header = tb.Frame(self.left_frame, bootstyle="primary", padding=12)
        header.pack(fill="x", pady=(0, 12))

        tb.Label(
            header,
            text="👤 Мій кабінет",
            font=("Segoe UI", 14, "bold"),
            bootstyle="inverse-primary"
        ).pack()

        # NAVIGATION (аналогично адмін-панелі)
        nav_frame = tb.Frame(self.left_frame, padding=5)
        nav_frame.pack(fill="x", pady=10)

        user_functions = [
            ("📊 Головна", "user_dashboard", "primary"),
            ("🚗 Мої авто", "my_purchases", "info"),
            ("📈 Статистика", "user_analytics", "warning"),
        ]

        for icon_text, destination, color in user_functions:
            tb.Button(
                nav_frame,
                text=icon_text,
                bootstyle=f"{color}",
                padding=8,
                command=lambda d=destination: self._user_navigate(d)
            ).pack(fill="x", pady=4)

        # (при желании можно добавить разделитель как в admin)
        # sep = ttk.Separator(nav_frame, orient="horizontal")
        # sep.pack(fill="x", pady=10)


    def _user_navigate(self, destination):
        """Навигация по пользовательской панели"""
        if destination == "user_dashboard":
            self._show_user_dashboard()
        elif destination == "my_purchases":
            self._show_my_purchases()
        elif destination == "user_analytics":
            self._show_user_analytics()

    def _show_user_dashboard(self):
        """Показать пользовательский дашборд"""
        self._clear_main_content()
        
        welcome_card = tb.Frame(self.main_content, bootstyle="info")
        welcome_card.pack(fill="x", pady=(0,15))
        
        tb.Label(welcome_card, text=f"Вітаємо, {self.current_user['username']}!", 
                font=("Segoe UI", 18, "bold"), bootstyle="inverse-info").pack(pady=12)

        stats_frame = tb.Frame(self.main_content)
        stats_frame.pack(fill="x", pady=8)

        try:
            cur = self.conn.cursor()
            
            cur.execute("SELECT COUNT(*) FROM purchases WHERE buyer_id = %s", (self.current_user['id'],))
            total_purchases = cur.fetchone()[0]
            
            cur.execute("SELECT COUNT(*) FROM purchases WHERE buyer_id = %s AND status_id in (1,2,3,4,5,6,7,8)", 
                       (self.current_user['id'],))
            active_deliveries = cur.fetchone()[0]
            
            cur.execute("SELECT COUNT(*) FROM purchases WHERE buyer_id = %s AND status_id = 9", 
                       (self.current_user['id'],))
            delivered = cur.fetchone()[0]
            
            cur.close()
            
            stats = [
                ("🚗 Всього авто", total_purchases, "primary"),
                ("📦 В дорозі", active_deliveries, "warning"),
                ("✅ Доставлено", delivered, "success"),
            ]
            
            for text, count, style in stats:
                card = tb.Frame(stats_frame, bootstyle=style, padding=8)
                card.pack(side="left", fill="x", expand=True, padx=3)
                tb.Label(card, text=text, font=("Segoe UI", 9), 
                        bootstyle=f"inverse-{style}").pack()
                tb.Label(card, text=str(count), font=("Segoe UI", 18, "bold"),
                        bootstyle=f"inverse-{style}").pack()
                
        except Error as e:
            print(f"Помилка завантаження статистики користувача: {e}")

        try:
            cur = self.conn.cursor(dictionary=True)
            cur.execute("""
                SELECT p.*, s.status_name,
                       DATEDIFF(CURDATE(), p.estimated_arrival_date) as days_late
                FROM purchases p
                LEFT JOIN statuses s ON p.status_id = s.status_id
                WHERE p.buyer_id = %s 
                AND p.estimated_arrival_date < CURDATE()
                AND p.is_delivered = FALSE
                AND s.status_name NOT LIKE '%Україні%'
                ORDER BY p.estimated_arrival_date ASC
            """, (self.current_user['id'],))
            late_cars = cur.fetchall()
            cur.close()

            if late_cars:
                late_frame = tb.LabelFrame(self.main_content, text="⚠️ Мої авто з запізненням", padding=12)
                late_frame.pack(fill="x", pady=15)
                
                late_cars_scroll_container = tb.Frame(late_frame)
                late_cars_scroll_container.pack(fill="both", expand=True)
                
                canvas = Canvas(late_cars_scroll_container, height=200)
                scrollbar = ttk.Scrollbar(late_cars_scroll_container, orient="vertical", command=canvas.yview)
                late_cars_content = tb.Frame(canvas)
                
                late_cars_content.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=late_cars_content, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)
                
                canvas.pack(side="left", fill="both", expand=True)
                scrollbar.pack(side="right", fill="y")
                
                for car in late_cars:
                    car_frame = tb.Frame(late_cars_content)
                    car_frame.pack(fill="x", pady=6)
                    
                    tb.Label(car_frame, 
                            text=f"{car['car_make']} {car['car_model']} ({car['car_year']})",
                            font=("Segoe UI", 11, "bold")).pack(anchor="w")
                    
                    tb.Label(car_frame, 
                            text=f"Запізнення: {car['days_late']} дн. • {car['status_name']}",
                            font=("Segoe UI", 9),
                            bootstyle="danger").pack(anchor="w")
                    
                    tb.Button(car_frame, text="👀 Деталі", bootstyle="outline",
                             command=lambda c=car: self._show_purchase_details(c)).pack(anchor="w", pady=3)
                    
                    ttk.Separator(car_frame, orient='horizontal').pack(fill='x', pady=3)
                    
        except Error as e:
            print(f"Помилка завантаження запізнілих авто для користувача: {e}")

    def _show_my_purchases(self):
        """Показать мои покупки (оновлений дизайн фільтрів)"""
        self._clear_main_content()
        
        header_frame = tb.Frame(self.main_content)
        header_frame.pack(fill="x", pady=(0, 8))
        
        tb.Label(
            header_frame,
            text="Мої автомобілі",
            font=("Segoe UI", 16, "bold")
        ).pack(side="left")
        
        # ---------- Картка фільтрів ----------
        filters_card = tb.LabelFrame(
            self.main_content,
            text="Фільтри",
            padding=10
        )
        filters_card.pack(fill="x", pady=5)

        # Перша лінія — статус + рік
        row1 = tb.Frame(filters_card)
        row1.pack(fill="x", pady=3)

        # Статус
        status_block = tb.Frame(row1, padding=(5, 2))
        status_block.pack(side="left", padx=5, fill="x", expand=True)

        tb.Label(
            status_block,
            text="⚙️ Статус",
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")

        self.user_status_filter = tb.StringVar(value="all")
        status_combo = tb.Combobox(
            status_block,
            textvariable=self.user_status_filter,
            values=["all", "bought_auction", "paid", "to_port", "at_port",
                    "in_sea", "in_klaipeda", "to_ukraine", "cleared_customs", "in_ukraine"],
            state="readonly",
            width=18
        )
        status_combo.pack(fill="x", pady=2)
        status_combo.bind("<<ComboboxSelected>>", self._load_my_purchases)

        # Рік
        year_block = tb.Frame(row1, padding=(5, 2))
        year_block.pack(side="left", padx=5, fill="x")

        tb.Label(
            year_block,
            text="📅 Рік",
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")

        self.user_year_filter = tb.StringVar(value="all")
        current_year = datetime.datetime.now().year
        years = [str(y) for y in range(current_year - 10, current_year + 1)]
        year_combo = tb.Combobox(
            year_block,
            textvariable=self.user_year_filter,
            values=["all"] + years,
            state="readonly",
            width=10
        )
        year_combo.pack(fill="x", pady=2)
        year_combo.bind("<<ComboboxSelected>>", self._load_my_purchases)

        # Друга лінія — пошук + кнопки
        row2 = tb.Frame(filters_card)
        row2.pack(fill="x", pady=5)
        
        self.user_search_var = tb.StringVar()
        search_entry = tb.Entry(
            row2,
            textvariable=self.user_search_var,
            width=40,
            font=("Segoe UI", 10)
        )
        search_entry.pack(side="left", padx=5, fill="x", expand=True)
        search_entry.insert(0, "Пошук по VIN, марці, моделі...")

        def on_search_focus_in(e):
            if search_entry.get() == "Пошук по VIN, марці, моделі...":
                search_entry.delete(0, 'end')
        
        def on_search_focus_out(e):
            if not search_entry.get():
                search_entry.insert(0, "Пошук по VIN, марці, моделі...")

        search_entry.bind("<FocusIn>", on_search_focus_in)
        search_entry.bind("<FocusOut>", on_search_focus_out)
        search_entry.bind("<KeyRelease>", lambda e: self._filter_my_purchases())

        tb.Button(
            row2,
            text="🔍 Пошук",
            bootstyle="info",
            command=self._filter_my_purchases
        ).pack(side="left", padx=5)

        tb.Button(
            row2,
            text="🗑️ Скинути",
            bootstyle="danger-outline",
            command=self._reset_user_filters
        ).pack(side="right", padx=5)
        
        # Контейнер карток
        cards_container = tb.Frame(self.main_content)
        cards_container.pack(fill="both", expand=True, pady=8)
        
        canvas = Canvas(cards_container)
        scrollbar = ttk.Scrollbar(cards_container, orient="vertical", command=canvas.yview)
        self.user_cards_frame = tb.Frame(canvas)
        
        self.user_cards_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.user_cards_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self._load_my_purchases()

    
    def _reset_user_filters(self):
        """Сброс пользовательских фильтров"""
        self.user_status_filter.set("all")
        self.user_year_filter.set("all")
        self.user_search_var.set("")
        self._load_my_purchases()
    
    def _load_my_purchases(self, event=None):
        """Загрузка моих покупок"""
        for w in self.user_cards_frame.winfo_children():
            w.destroy()
        
        try:
            cur = self.conn.cursor(dictionary=True)
            
            query = """
                SELECT p.*, c.country_name, a.auction_name, l.location_name, 
                       s.status_name, u.username, port.port_name,
                       EXISTS(SELECT 1 FROM purchase_images WHERE purchase_id = p.purchase_id) as has_images,
                       (SELECT image_url FROM purchase_images WHERE purchase_id = p.purchase_id LIMIT 1) as first_image_path
                FROM purchases p
                LEFT JOIN countries c ON p.country_id = c.country_id
                LEFT JOIN auctions a ON p.auction_id = a.auction_id
                LEFT JOIN locations l ON p.location_id = l.location_id
                LEFT JOIN statuses s ON p.status_id = s.status_id
                LEFT JOIN users u ON p.buyer_id = u.id
                LEFT JOIN locations loc ON p.location_id = loc.location_id
                LEFT JOIN ports port ON loc.default_port_id = port.port_id
                WHERE p.buyer_id = %s
            """
            
            params = [self.current_user['id']]
            
            status_filter = self.user_status_filter.get()
            if status_filter != "all":
                query += " AND s.status_key = %s"
                params.append(status_filter)
            
            year_filter = self.user_year_filter.get()
            if year_filter != "all":
                query += " AND p.car_year = %s"
                params.append(int(year_filter))
            
            query += " ORDER BY p.purchase_date DESC"
            
            cur.execute(query, params)
            purchases = cur.fetchall()
            cur.close()
            
            row_frame = None
            for idx, purchase in enumerate(purchases):
                if idx % 3 == 0:
                    row_frame = tb.Frame(self.user_cards_frame)
                    row_frame.pack(fill="x", pady=3)
                
                card = CarCard(row_frame, purchase, on_click=self._show_purchase_details)
                card.pack(side="left", padx=8, fill="both", expand=True)
            
            if not purchases:
                tb.Label(self.user_cards_frame, text="У вас ще немає покупок за обраними фільтрами", 
                        font=("Segoe UI", 12)).pack(pady=40)
                
        except Error as e:
            messagebox.showerror("Помилка", f"Помилка завантаження: {e}")
    
    def _filter_my_purchases(self):
        """Фильтрация моих покупок"""
        search_text = self.user_search_var.get().lower()
        if search_text == "Пошук по VIN, марці, моделі...":
            search_text = ""
        
        for w in self.user_cards_frame.winfo_children():
            w.destroy()
        
        try:
            cur = self.conn.cursor(dictionary=True)
            
            query = """
                SELECT p.*, c.country_name, a.auction_name, l.location_name, 
                       s.status_name, u.username, port.port_name,
                       EXISTS(SELECT 1 FROM purchase_images WHERE purchase_id = p.purchase_id) as has_images,
                       (SELECT image_url FROM purchase_images WHERE purchase_id = p.purchase_id LIMIT 1) as first_image_path
                FROM purchases p
                LEFT JOIN countries c ON p.country_id = c.country_id
                LEFT JOIN auctions a ON p.auction_id = a.auction_id
                LEFT JOIN locations l ON p.location_id = l.location_id
                LEFT JOIN statuses s ON p.status_id = s.status_id
                LEFT JOIN users u ON p.buyer_id = u.id
                LEFT JOIN locations loc ON p.location_id = loc.location_id
                LEFT JOIN ports port ON loc.default_port_id = port.port_id
                WHERE p.buyer_id = %s
            """
            
            params = [self.current_user['id']]
            
            status_filter = self.user_status_filter.get()
            if status_filter != "all":
                query += " AND s.status_key = %s"
                params.append(status_filter)
            
            year_filter = self.user_year_filter.get()
            if year_filter != "all":
                query += " AND p.car_year = %s"
                params.append(int(year_filter))
            
            if search_text:
                query += " AND (LOWER(p.vin_number) LIKE %s OR LOWER(p.car_make) LIKE %s OR LOWER(p.car_model) LIKE %s)"
                params.extend([f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"])
            
            query += " ORDER BY p.purchase_date DESC"
            
            cur.execute(query, params)
            purchases = cur.fetchall()
            cur.close()
            
            row_frame = None
            for idx, purchase in enumerate(purchases):
                if idx % 3 == 0:
                    row_frame = tb.Frame(self.user_cards_frame)
                    row_frame.pack(fill="x", pady=3)
                
                card = CarCard(row_frame, purchase, on_click=self._show_purchase_details)
                card.pack(side="left", padx=8, fill="both", expand=True)
            
            if not purchases:
                tb.Label(self.user_cards_frame, text="Нічого не знайдено", 
                        font=("Segoe UI", 12)).pack(pady=40)
                
        except Error as e:
            messagebox.showerror("Помилка", f"Помилка пошуку: {e}")

    def _show_user_analytics(self):
        """Показати персональну аналітику користувача"""
        self._clear_main_content()

        analytics_frame = tb.Frame(self.main_content)
        analytics_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # ───── Заголовок + кнопка звіту ─────────────────────────
        header_frame = tb.Frame(analytics_frame)
        header_frame.pack(fill="x", pady=(0, 10))

        tb.Label(
            header_frame,
            text="📈 Моя статистика",
            font=("Segoe UI", 18, "bold")
        ).pack(side="left")

        tb.Button(
            header_frame,
            text="📄 Мій звіт",
            bootstyle="info",
            padding=5,
            command=self._open_user_report_dialog
        ).pack(side="right")

        # ───── SQL аналітика ─────────────────────────
        try:
            cur = self.conn.cursor()
            cur.execute("""
                SELECT s.status_name, COUNT(p.purchase_id) AS count
                FROM statuses s
                LEFT JOIN purchases p 
                    ON s.status_id = p.status_id 
                    AND p.buyer_id = %s
                GROUP BY s.status_id, s.status_name
                ORDER BY s.order_index
            """, (self.current_user["id"],))

            status_stats = cur.fetchall()
            cur.close()

            stats_frame = tb.Frame(analytics_frame)
            stats_frame.pack(fill="x", pady=10)

            # Максимум потрібен для нормальної шкали
            max_value = max((c for (_, c) in status_stats), default=1)

            LABEL_WIDTH = 28
            NUMBER_WIDTH = 4

            # ───── Вивід рядків аналітики ─────────────────────────
            for status_name, count in status_stats:
                row = tb.Frame(stats_frame)
                row.pack(fill="x", pady=4)

                name_lower = status_name.lower()
                if "куп" in name_lower:
                    style = "info"
                elif "порт" in name_lower:
                    style = "warning"
                elif "мор" in name_lower:
                    style = "primary"
                elif "укра" in name_lower:
                    style = "success"
                else:
                    style = "secondary"

                tb.Label(
                    row,
                    text=status_name,
                    width=LABEL_WIDTH,
                    anchor="w",
                    font=("Segoe UI", 11)
                ).pack(side="left")

                percent = int((count / max_value) * 100) if max_value else 0

                pb = tb.Progressbar(
                    row,
                    value=percent,
                    maximum=100,
                    bootstyle=style
                )
                pb.pack(side="left", fill="x", expand=True, padx=10, ipady=2)

                tb.Label(
                    row,
                    text=str(count),
                    width=NUMBER_WIDTH,
                    anchor="e",
                    font=("Segoe UI", 11, "bold")
                ).pack(side="right")

        except Exception as e:
            tb.Label(
                analytics_frame,
                text=f"Помилка завантаження статистики:\n{e}",
                bootstyle="danger"
            ).pack()

        # ───── Кнопка назад ─────────────────────────
        tb.Button(
            analytics_frame,
            text="← Назад",
            bootstyle="dark",
            padding=5,
            command=self._show_user_dashboard
        ).pack(side="bottom", pady=10)

    def _open_user_report_dialog(self):
        dlg = tb.Toplevel(self)
        dlg.title("Мій звіт за період")
        dlg.geometry("420x300")
        dlg.transient(self)
        dlg.grab_set()
        self.center_window(dlg, 420, 370)

        selected = {"from": None, "to": None, "file": None}

        card = tb.Frame(dlg, padding=15, borderwidth=1, relief="solid")
        card.pack(fill="both", expand=True, padx=15, pady=15)

        tb.Label(card, text="📅 Оберіть діапазон дат", font=("Segoe UI", 14, "bold")).pack(pady=(0, 10))

        # --- From ---
        frm_section = tb.Frame(card)
        frm_section.pack(fill="x", pady=6)

        tb.Label(frm_section, text="З дати:", font=("Segoe UI", 11)).pack(anchor="w")

        from_btn = tb.Button(frm_section, text="Не обрано",
                            bootstyle="secondary-outline", width=26)
        from_btn.pack(pady=3)

        def pick_from():
            cal = CalendarDialog(self)
            self.wait_window(cal.dialog)
            if cal.result:
                selected["from"] = cal.result
                from_btn.configure(text=str(cal.result))

        from_btn.configure(command=pick_from)

        # --- To ---
        to_section = tb.Frame(card)
        to_section.pack(fill="x", pady=6)

        tb.Label(to_section, text="По дату:", font=("Segoe UI", 11)).pack(anchor="w")

        to_btn = tb.Button(to_section, text="Не обрано",
                        bootstyle="secondary-outline", width=26)
        to_btn.pack(pady=3)

        def pick_to():
            cal = CalendarDialog(self)
            self.wait_window(cal.dialog)
            if cal.result:
                selected["to"] = cal.result
                to_btn.configure(text=str(cal.result))

        to_btn.configure(command=pick_to)

        # --- SAVE PATH ---
        tb.Label(card, text="Файл звіту:", font=("Segoe UI", 11)).pack(pady=(10, 3))

        file_btn = tb.Button(card, text="📁 Обрати місце збереження", bootstyle="info")
        file_btn.pack()

        def pick_file():
            file = filedialog.asksaveasfilename(
                title="Зберегти звіт",
                defaultextension=".csv",
                filetypes=[
                    ("CSV файл", "*.csv"),
                    ("Excel файл", "*.xlsx"),
                    ("Всі файли", "*.*")
                ]
            )
            if file:
                selected["file"] = file
                file_btn.configure(text=file)

        file_btn.configure(command=pick_file)

        # --- GENERATE ---
        def generate():
            if not selected["from"] or not selected["to"]:
                messagebox.showwarning("Помилка", "Оберіть обидві дати.")
                return

            if selected["from"] > selected["to"]:
                messagebox.showwarning("Помилка", "Дата 'З' не може бути більшою за 'По'.")
                return

            if not selected["file"]:
                messagebox.showwarning("Помилка", "Оберіть шлях збереження звіту.")
                return

            dlg.destroy()

            # сохраняем диапазон для отображения в аналитике
            self.user_report_range = {
                "from": str(selected["from"]),
                "to": str(selected["to"])
            }

            self._generate_user_report(selected)
            self._show_user_analytics()

        tb.Button(
            card,
            text="📄 Сформувати звіт",
            bootstyle="success",
            width=20,
            command=generate
        ).pack(pady=20)

    def _generate_user_report(self, data):
        date_from = str(data["from"])
        date_to = str(data["to"])
        file = data["file"]
        ext = os.path.splitext(file)[1].lower()

        try:
            # --- 1. Берём ВСЕ колонки purchases ---
            cur = self.conn.cursor()
            cur.execute("SHOW COLUMNS FROM purchases")
            columns = [col[0] for col in cur.fetchall()]
            cur.close()

            # --- 2. Делаем запрос только для текущего пользователя ---
            query = f"""
                SELECT {", ".join(columns)}
                FROM purchases
                WHERE buyer_id = %s
                AND DATE(purchase_date) BETWEEN %s AND %s
                ORDER BY purchase_date ASC
            """

            cur = self.conn.cursor(dictionary=True)
            cur.execute(query, (self.current_user["id"], date_from, date_to))
            rows = cur.fetchall()
            cur.close()

            total = len(rows)

            if not rows:
                messagebox.showinfo("Звіт", "Немає ваших покупок за даний період.")
                return

            # --- CSV ---
            if ext == ".csv":
                import csv
                with open(file, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=columns)
                    writer.writeheader()
                    writer.writerows(rows)

                    f.write("\n")
                    f.write(
                        f"За період з {date_from} по {date_to} ви купили {total} автомобілів.\n"
                    )

            # --- Excel ---
            elif ext == ".xlsx":
                import pandas as pd
                from openpyxl import load_workbook

                df_pd = pd.DataFrame(rows)
                df_pd.to_excel(file, index=False)

                wb = load_workbook(file)
                ws = wb.active
                ws.append([])
                ws.append([f"За період з {date_from} по {date_to} ви купили {total} автомобілів."])
                wb.save(file)

            messagebox.showinfo("Готово", f"Звіт збережено:\n{file}")

        except Exception as e:
            import traceback
            messagebox.showerror("Помилка", traceback.format_exc())


    def _toggle_theme(self):
        """Переключение темы"""
        self.dark_mode = not self.dark_mode
        if self.dark_mode:
            self.style.theme_use("darkly")
        else:
            self.style.theme_use("flatly")
        
        self._build_main_ui()

    def _logout(self):
        """Выход из системы"""
        try:
            if self.conn:
                self.conn.close()
        except:
            pass
        self.current_user = None
        self.conn = None
        self._build_login_ui()